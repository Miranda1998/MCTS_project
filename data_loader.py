# data_loader.py
from openpyxl import load_workbook
import numpy as np
import os
from args import parse_args
import json
import pickle

# ===== 你原来的 Load_Data：不要删，保持兼容 =====
def Load_Data(args, vessels_num):
    # （原文件内容保持不变）
    args = parse_args()
    ScheduleHorizon = args.ScheduleHorizon
    slot_gap = args.slot_gap
    T = int(ScheduleHorizon * (60 / slot_gap))

    drone_battery = args.drone_battery
    drone_base_num = args.drone_base_num

    drone_speed = args.drone_speed
    drone_obs_dwell = args.drone_obs_dwell

    drone_feasible_path = os.path.join(args.rouFeasiblePairs,
                                       f"new_feasible_pair20260223_TASK={vessels_num}_T={T}.npy")
    drone_feasible_pair = np.load(drone_feasible_path, allow_pickle=True)


    tpb = load_workbook(filename=args.rouReward)
    tps = tpb.worksheets[0]
    reward = []
    for row_A in range(1, vessels_num + 1):
        a1 = tps.cell(row=row_A, column=1).value
        if a1:
            reward.append(a1)

    drone_feasible_pair = set(map(tuple, drone_feasible_pair.tolist()))
    return (T, slot_gap, drone_battery, drone_base_num, drone_feasible_pair, reward, drone_speed, drone_obs_dwell)

# ===== 新增：第四个工作 MCTS 在线评估所需轨迹加载 =====
def Load_Trajectories(args, n_vessels=None):
    """
    Returns:
      traj_true:  [N, K, 2]  (km or same coordinate system)
      traj_mean:  [N, K, 2]
      scenarios:  [M, N, K, 2]
    """
    traj_true = np.load(args.rouTrajTrue, allow_pickle=True)
    traj_mean = np.load(args.rouTrajMean, allow_pickle=True)
    scenarios = np.load(args.rouScenarios, allow_pickle=True)

    # 若提供 n_vessels，则截取前 n_vessels
    if n_vessels is not None:
        traj_true = traj_true[:n_vessels]
        traj_mean = traj_mean[:n_vessels]
        scenarios = scenarios[:, :n_vessels]

    # 基本形状检查
    if traj_true.ndim != 3 or traj_true.shape[-1] != 2:
        raise ValueError(f"traj_true shape expected [N,K,2], got {traj_true.shape}")
    if traj_mean.shape != traj_true.shape:
        raise ValueError(f"traj_mean must match traj_true shape, got {traj_mean.shape} vs {traj_true.shape}")
    if scenarios.ndim != 4 or scenarios.shape[1:] != traj_true.shape:
        raise ValueError(f"scenarios shape expected [M,N,K,2] = [M,{traj_true.shape[0]},{traj_true.shape[1]},2], got {scenarios.shape}")

    return traj_true.astype(float), traj_mean.astype(float), scenarios.astype(float)