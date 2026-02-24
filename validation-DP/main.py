from datetime import datetime
import os
import json
import time
from openpyxl import Workbook
from contextlib import redirect_stdout
from args import parse_args
from data_loader import Load_Data
from ADMM import *
import sys

class Tee:
    def __init__(self, *streams):
        self.streams = streams

    def write(self, data):
        for s in self.streams:
            s.write(data)

    def flush(self):
        for s in self.streams:
            s.flush()


def _ensure_header(ws):
    if ws.max_row == 1 and ws["A1"].value is None:
        ws.append([
            "平台规模(AEOS,Airship,UAV)",
            "任务规模(船舶数)",
            "final_UB",
            "final_LB",
            "final_GAP",
            "用时(s)"
        ])

def run_task_group(args, AEOSs, airships, drones, vessels_num, wb, rou0, date_today):
    ws_name = "ADMM_DP_result"
    ws = wb[ws_name] if ws_name in wb.sheetnames else wb.create_sheet(ws_name)
    _ensure_header(ws)

    print('当前平台数量为:', (AEOSs, airships, drones))
    print('当前船舶数量为:', vessels_num)

    result_dir = os.path.join(rou0, 'result', date_today)
    os.makedirs(result_dir, exist_ok=True)

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    t0 = time.perf_counter()

    (T, slot_gap, drone_battery, drone_base_num, drone_feasible_pair, reward) \
        = Load_Data(args, vessels_num, AEOSs, airships, drones)

    params = ADMMParams(
        lambda_=args.lambda_,
        rho=args.rho,
        max_iter=args.max_iter,
        tol=args.tol,
        seed=args.seed,
        gap_tol=args.MIPgap,
        time_limit=args.Timelimit
    )


    inst = CSHEOSPInstance(
        T=T,
        slot_gap=slot_gap,
        drone_battery=drone_battery,
        drone_base_num=drone_base_num,
        vessels_num=vessels_num,
        reward=reward,
        uav_arcs=drone_feasible_pair
    )

    # 单线程版本（串行高斯赛德尔）
    res = solve_admm_dp_serial_gs(inst, AEOSs, airships, drones, params, t0)

    print(f"[OK] final_best_obj= { res.best_obj}")
    print(f"[OK] final_best_sol= { res.best_sol}")


def main():
    args = parse_args()
    rou0 = os.getcwd()
    wb = Workbook()

    # remove default empty sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])

    date_now = "2026_0220_admm_dp_serial_gs_final"

    result_dir = os.path.join(rou0, 'result', date_now)
    os.makedirs(result_dir, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    txt_path = os.path.join(result_dir, f"exp_{stamp}_ADMM_DP.log")
    xlsx_path = os.path.join(result_dir, f"exp_{stamp}_ADMM_DP.xlsx")

    old_stdout = sys.stdout
    with open(txt_path, "w", encoding="utf-8") as f:
        sys.stdout = Tee(old_stdout, f)  # 同时输出到控制台 + 文件
        try:
            for its in range(2, 3, 1):
                AEOSs, airships, drones = 0, 0, its

                for vessels_num in range(10, 60, 20):
                    print(f"\n=== SPs={(AEOSs, airships, drones)}, vessels_num={vessels_num}====")
                    run_task_group(args, AEOSs, airships, drones, vessels_num, wb, rou0, date_today=date_now)

                    wb.save(xlsx_path)
                    print(f"[OK] wrote excel: {xlsx_path}")
                    print(f"[OK] wrote log:   {txt_path}")
        finally:
            sys.stdout = old_stdout
            wb.save(xlsx_path)


if __name__ == '__main__':
    main()
