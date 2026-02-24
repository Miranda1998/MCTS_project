from openpyxl import load_workbook
import pandas as pd
import numpy as np
import math
import os
import json
import pickle

# 读取返回的各个变量，并根据变量设置集合等参数
def Load_Data(args, vessels_num, AEOS_num, airships_num, drones_num):

    # 和问题有关的参数
    with open(args.rouSPs, 'r') as f:
        config = json.load(f)

    # 调度周期参数
    ScheduleHorizon = config['ScheduleHorizon']
    slot_gap = config['SlotGap']
    T = int(ScheduleHorizon * (60 / slot_gap))

    # 无人机相关参数
    drone_battery = config['drone_battery']
    drone_base_num = config['drone_base_num']


    # feasible_pair相关参数
    # 无人机feasible_pair
    drone_feasible_path = os.path.join(args.rouFeasiblePairs,
                                       f"new_feasible_pair20260223_TASK={vessels_num}_T={T}.npy")
    drone_feasible_pair = np.load(drone_feasible_path, allow_pickle=True)


    # 奖励相关参数
    tpb = load_workbook(filename=args.rouReward)
    tps = tpb.worksheets[0]
    reward = []
    for row_A in range(1, vessels_num + 1):
        a1 = tps.cell(row=row_A, column=1).value
        if a1:
            reward.append(a1)


    drone_feasible_pair = set(map(tuple, drone_feasible_pair.tolist()))


    return (T, slot_gap, drone_battery, drone_base_num, drone_feasible_pair, reward)


