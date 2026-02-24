# main.py
from datetime import datetime
import os
import time
import numpy as np
import sys
from openpyxl import Workbook
from args import parse_args
from data_loader import Load_Data, Load_Trajectories

# ===== 你的 ADMM 相关仍可保留 =====

from mcts.env import OnlineEnv, OnlineEnvConfig
from mcts.mcts_core import MCTSPlanner
from mcts.scenario_updater import ScenarioUpdater
from mcts.evaluator import run_episode

class Tee:
    def __init__(self, *streams):
        self.streams = streams
    def write(self, data):
        for s in self.streams:
            s.write(data)
    def flush(self):
        for s in self.streams:
            s.flush()

def _ensure_header_mcts(ws):
    if ws.max_row == 1 and ws["A1"].value is None:
        ws.append([
            "exp_name",
            "rollout_policy",
            "N(vessels)",
            "K(steps)",
            "M(scenarios)",
            "uav_num",
            "tau(min)",
            "budget(ms)",
            "repeat_id",
            "total_reward",
            "served_cnt",
            "decision_cnt",
            "walltime(s)"
        ])

def run_mcts_group(args, wb, vessels_num, drones_num):
    ws_name = "MCTS_result"
    ws = wb[ws_name] if ws_name in wb.sheetnames else wb.create_sheet(ws_name)
    _ensure_header_mcts(ws)

    # load trajectories
    traj_true, traj_mean, scenarios = Load_Trajectories(args, n_vessels=vessels_num)
    N, K, _ = traj_true.shape
    M = scenarios.shape[0]

    (horizon_steps, slot_gap, drone_battery, drone_base_num, drone_feasible_pair, reward,
     drone_speed, drone_obs_dwell) = Load_Data(args, vessels_num)

    info_delay_steps = int(round(args.info_delay_min / slot_gap))

    print('info_delay_steps', info_delay_steps)

    cfg = OnlineEnvConfig(
        slot_gap_min=float(args.slot_gap),
        horizon_steps=int(horizon_steps),
        info_delay_steps=int(info_delay_steps),
        vessels_num=int(vessels_num),
        uav_num=int(drones_num),
        uav_battery_capacity=int(drone_battery),
        uav_speed_kmph=float(args.drone_speed),
        feasible_pair=drone_feasible_pair,
        uav_obs_dwell_steps=int(drone_obs_dwell),
        hit_radius_km_aeos=float(args.hit_radius_km_aeos),
        hit_radius_km_airship=float(args.hit_radius_km_airship),
        hit_radius_km_drones=float(args.hit_radius_km_drones),
        base_set=list(range(int(drone_base_num)))
    )

    env = OnlineEnv(
        traj_true=traj_true,
        traj_mean=traj_mean,
        scenarios=scenarios,
        reward=reward,
        cfg=cfg,
        normal_plan=None  # TODO: 把 ADMM-LNS 输出的常态化监测计划塞进来
    )

    # 观测一致性门限：简单取 AEOS 半径或其比例
    gate = float(args.hit_radius_km_aeos)
    updater = ScenarioUpdater(scenarios=scenarios, gate_km=gate)

    print(f"[MCTS] N={N}, K={horizon_steps}, M={M}, uav={drones_num}, tau={args.info_delay_min}min")

    for rep in range(args.repeat):
        t0 = time.perf_counter()
        planner = MCTSPlanner(
            env=env,
            ucb_c=float(args.ucb_c),
            rollout_depth=int(args.rollout_depth),
            rollout_policy=str(args.rollout_policy),
            seed=int(args.seed) + rep
        )
        res = run_episode(
            env=env,
            planner=planner,
            updater=updater,
            budget_ms=int(args.mcts_budget_ms),
            iters=int(args.mcts_iters)
        )
        t1 = time.perf_counter()
        ws.append([
            str(args.exp_name),
            str(args.rollout_policy),
            int(N),
            int(horizon_steps),
            int(M),
            int(drones_num),
            float(args.info_delay_min),
            int(args.mcts_budget_ms if args.mcts_iters == 0 else args.mcts_iters),
            int(rep),
            float(res.total_reward),
            int(res.served_cnt),
            int(res.decision_cnt),
            float(t1 - t0)
        ])
        print(f"[rep={rep}] reward={res.total_reward:.4f}, served={res.served_cnt}, decisions={res.decision_cnt}, time={t1-t0:.3f}s")

def main():
    args = parse_args()
    rou0 = os.getcwd()
    wb = Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])

    date_now = datetime.now().strftime("%Y_%m%d_%H%M") + "_" + str(args.exp_name)
    result_dir = os.path.join(rou0, 'result', date_now)
    os.makedirs(result_dir, exist_ok=True)

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')


    for rollout_policy in ["random", "eps_greedy", "adp"]:
        args.rollout_policy = rollout_policy

        txt_path = os.path.join(result_dir, f"exp_{stamp}_{args.rollout_policy}.log")
        xlsx_path = os.path.join(result_dir, f"exp_{stamp}_{args.rollout_policy}.xlsx")

        old_stdout = sys.stdout
        with open(txt_path, "w", encoding="utf-8") as f:
            sys.stdout = Tee(old_stdout, f)
            try:
                for drones in range(1, 4):
                    for vessels in range(10, 60, 20):
                        for time_budget in [5000, 10000, 15000, 20000, 25000, 30000]:
                            args.mcts_budget_ms = time_budget
                            print(f"\n=== SPs={(4, 4, drones)}, vessels_num={vessels}, rollout={args.rollout_policy}====")
                            print(f"\n=== time_budget={time_budget}")
                            run_mcts_group(args, wb, vessels_num=vessels, drones_num=drones)
                wb.save(xlsx_path)
                print(f"[OK] wrote excel: {xlsx_path}")
                print(f"[OK] wrote log:   {txt_path}")
            finally:
                sys.stdout = old_stdout
                wb.save(xlsx_path)

if __name__ == '__main__':
    main()